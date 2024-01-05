import argparse
import concurrent.futures
import json
from datetime import datetime

import openpyxl
import stripe
from dotenv import dotenv_values

# Load variables from the .env file
env_variables = dotenv_values(".env")
stripe.api_key = env_variables["STRIPE_API_KEY"]

# Product name replacement string if the intent is a direct payment and has no defined product.
DIRECT_CHARGE = "SudoSOS Topup"


def get_payout_transactions(balance_payout_id):
    """
    Retrieve transactions associated with a balance payout ID.
    """
    transactions = stripe.BalanceTransaction.auto_paging_iter(
        payout=balance_payout_id,
        expand=["data.source.payment_intent"],
    )
    # The transactions also include the payout to bank, so we filter on type.
    transactions_data = [transaction for transaction in transactions if transaction.type != "payout"]
    return transactions_data


def get_payment_intents_ids(transactions_data):
    """
    Extract payment intent IDs from transactions data.
    """
    payment_intents_ids = [transaction.source.payment_intent.id for transaction in transactions_data if
                           hasattr(transaction.source, 'payment_intent')]
    return payment_intents_ids


def make_intent_charge_dict(payment_intents_ids):
    """
    Create a dictionary mapping payment intent IDs to product and customer.

    We do this concurrently since we need to make an API call for each transaction.
    """
    intent_type_charge_dict = {}
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = []
        for intent_id in payment_intents_ids:
            futures.append(executor.submit(get_intent_data, intent_id))

        for future in concurrent.futures.as_completed(futures):
            intent_id, intent_data = future.result()
            intent_type_charge_dict[intent_id] = intent_data

    return intent_type_charge_dict


def get_intent_data(intent_id):
    charges = stripe.checkout.Session.list(payment_intent=intent_id, expand=["data.line_items"])

    if len(charges) > 1:
        raise ValueError('Intent has more than 1 session.')

    if len(charges) == 0:
        return intent_id, {'product': DIRECT_CHARGE, 'customer': {'name': None, 'email': None}}

    for charge in charges:
        return intent_id, {'product': charge.line_items.data[0].description, 'customer': charge.customer_details}


def simplify_report_data(transactions_data, intent_type_charge_dict):
    """
    Simplify report data by extracting the required fields from transactions_data
    and intent_type_charge_dict.
    """
    simplified_data = []
    for transaction in transactions_data:
        payment_intent = transaction.source.payment_intent.id
        intent_data = intent_type_charge_dict.get(payment_intent,
                                                  {"product": None, "customer": {"name": None, "email": None}})
        simplified_transaction = {
            "id": transaction.id,
            "created": transaction.created,
            "amount": transaction.amount,
            "currency": transaction.currency,
            "fee": transaction.fee,
            "net": transaction.net,
            "product": intent_data["product"],
            "name": intent_data["customer"]["name"],
            "email": intent_data["customer"]["email"],
        }
        simplified_data.append(simplified_transaction)
    return simplified_data


def aggregate_report_transactions(report_transactions):
    """
    Aggregate report transactions by summing amount, fee, and net for each unique product.
    """
    aggregated_data = {}

    for transaction in report_transactions:
        product = transaction["product"]
        amount = transaction["amount"]
        fee = transaction["fee"]
        net = transaction["net"]

        if product not in aggregated_data:
            aggregated_data[product] = {
                "amount": amount,
                "fee": fee,
                "net": net,
                "product": product
            }
        else:
            aggregated_data[product]["amount"] += amount
            aggregated_data[product]["fee"] += fee
            aggregated_data[product]["net"] += net

    return aggregated_data


def get_report_data(balance_payout_id):
    """
      Retrieve and combine all data needed for generating a report.
    """
    transactions_data = get_payout_transactions(balance_payout_id)
    payment_intents_ids = get_payment_intents_ids(transactions_data)
    intent_type_charge_dict = make_intent_charge_dict(payment_intents_ids)

    report_transactions = simplify_report_data(transactions_data, intent_type_charge_dict)
    aggregation = aggregate_report_transactions(report_transactions)
    report_data = {"balance_payout_id": balance_payout_id,
                   "transactions": report_transactions,
                   "aggregation": aggregation}

    return report_data


def read_report_data_from_json(file_path):
    with open(file_path, 'r') as file:
        report_data = json.load(file)

    return report_data


def format_for_worksheet(data, key):
    formatted = data[key]
    if key in ["amount", "fee", "net"]:
        formatted = data[key] / 100
    elif key == "created":
        formatted = datetime.utcfromtimestamp(data[key]).strftime("%Y-%m-%d %H:%M:%S")
    return formatted


def save_to_worksheet(report_data, name='output.xlsx'):
    headers = ["id", "created", "amount", "currency", "fee", "net", "product", "name", "email"]
    numeric = ["amount", "fee", "net"]

    aggregation_headers = ["product"] + numeric
    aggregation_len = len(report_data["aggregation"]) + 2

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for idx, header in enumerate(aggregation_headers):
        worksheet.cell(row=1, column=idx + 1, value=header)

    for row_idx, data in enumerate(report_data["aggregation"]):
        for col_idx, key in enumerate(aggregation_headers):
            cell = worksheet.cell(row=row_idx + 2, column=col_idx + 1)
            cell.value = format_for_worksheet(report_data["aggregation"][data], key)

    for idx, header in enumerate(headers):
        worksheet.cell(row=aggregation_len + 1, column=idx + 1, value=header)

    for row_idx, data in enumerate(report_data["transactions"]):
        for col_idx, key in enumerate(headers):
            cell = worksheet.cell(row=row_idx + 2 + aggregation_len, column=col_idx + 1)
            cell.value = format_for_worksheet(data, key)

    workbook.save(name)


def make_report(balance_payout_id, name=''):
    report_data = get_report_data(balance_payout_id)
    save_to_worksheet(report_data, name)


def get_latest_payouts(args):
    n = args.poll_stripe
    payouts = stripe.Payout.list(limit=n)
    for i, payout in enumerate(payouts.data, start=1):
        print(
            f"{i}. Created: {datetime.utcfromtimestamp(payout.created).strftime('%Y-%m-%d')}, ID: {payout.id}")

    if n > 0:
        while True:
            try:
                choice = int(input("Enter the number of the payout to generate a report (0 to exit): "))
                if choice == 0:
                    break
                if choice > n or choice < 0:
                    print("Invalid choice. Please enter a valid number.")
                    continue
                args.payout = payouts.data[choice - 1].id
                process_report_data_from_args(args)
                break
            except ValueError:
                print("Invalid input. Please enter a valid number.")


def parse_arguments():
    parser = argparse.ArgumentParser(description="Process report data.")
    parser.add_argument("-s", "--poll-stripe", metavar="N", type=int,
                        help="Fetch the latest N BalancePayments and print their age and ID")
    parser.add_argument("-p", "--payout", help="Payout ID")
    parser.add_argument("-j", "--json", help="Path to JSON file")
    parser.add_argument("-c", "--print-json", action="store_true", help="Print JSON result, default is false.")
    parser.add_argument("-x", "--save-to-excel", action="store_false", help="Save to Excel file, default is true.")
    parser.add_argument("-o", "--output-name", help="Output file name for Excel")
    return parser.parse_args()


def process_report_data_from_args(args):
    payout_id = args.payout
    json_file_path = args.json
    print_json = args.print_json
    save_to_excel = args.save_to_excel
    output_name = args.output_name
    process_report_data(payout_id, json_file_path, print_json, save_to_excel, output_name)


def process_report_data(payout_id, json_file_path=None, print_json=False, save_to_excel=True, output_name=None):
    if json_file_path:
        report_data = read_report_data_from_json(json_file_path)
    elif payout_id:
        report_data = get_report_data(payout_id)
    else:
        raise ValueError("Either payout ID or JSON file must be provided.")

    if print_json:
        print(json.dumps(report_data, indent=4))

    if save_to_excel:
        if output_name is None:
            output_name = f"Report {report_data['balance_payout_id']}.xlsx"
        save_to_worksheet(report_data, name=output_name)


def main():
    if not stripe.api_key.startsWith('rk_'):
        raise ValueError("STRIPE_API_KEY must be set to a restricted API key. Check if the environment variable has been set and whether the key is a restricted API key")


    args = parse_arguments()

    if args.poll_stripe is not None:
        get_latest_payouts(args)
    else:
        process_report_data_from_args(args)


if __name__ == "__main__":
    main()
